"""
main.py
=======
The single entry point for the entire LBO analysis.

Run this file to:
  1. Load deal assumptions
  2. Project 5-year financials
  3. Build debt repayment schedule
  4. Calculate IRR, MoM, return attribution
  5. Run Bull / Base / Bear / Distressed scenarios
  6. Generate dual sensitivity heatmaps
  7. Export a formatted Excel tearsheet
  8. Save all charts to /outputs/charts/

Usage:
  python main.py                     # runs base case (Devyani International)
  python main.py --entry 7.5         # override entry multiple
  python main.py --exit 10           # override exit multiple
  python main.py --cagr 0.15         # override revenue growth
  python main.py --company "Sapphire Foods"  # change company name

Everything is modular — swap in any company's numbers via assumptions.py
and this entire pipeline re-runs automatically.
"""

import os
import sys
import argparse
import warnings
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import numpy as np

warnings.filterwarnings("ignore")

# ── Local module imports ───────────────────────────────────
from assumptions      import DealAssumptions
from income_statement import project_income_statement, print_income_statement
from debt_schedule    import build_debt_schedule, print_debt_schedule
from returns          import calculate_returns, print_returns
from scenarios        import run_all_scenarios, build_scenario_table, plot_scenario_comparison
from sensitivity      import run_all_sensitivities


# ══════════════════════════════════════════════════════════
#  SECTION 1: ARGUMENT PARSER
#  Lets you override assumptions from the command line
#  without touching assumptions.py
# ══════════════════════════════════════════════════════════

def parse_args():
    parser = argparse.ArgumentParser(
        description="LBO Financial Model — Python"
    )
    parser.add_argument("--company", type=str,   default=None)
    parser.add_argument("--entry",   type=float, default=None, help="Entry EV/EBITDA multiple")
    parser.add_argument("--exit",    type=float, default=None, help="Exit EV/EBITDA multiple")
    parser.add_argument("--cagr",    type=float, default=None, help="Revenue CAGR (e.g. 0.12)")
    parser.add_argument("--debt",    type=float, default=None, help="Debt % of EV (e.g. 0.65)")
    parser.add_argument("--hold",    type=int,   default=None, help="Holding period in years")
    parser.add_argument("--no-plots",action="store_true",      help="Skip chart generation")
    return parser.parse_args()


# ══════════════════════════════════════════════════════════
#  SECTION 2: DIRECTORY SETUP
# ══════════════════════════════════════════════════════════

def setup_directories():
    dirs = [
        "outputs/charts",
        "outputs/tearsheet",
        "data",
    ]
    for d in dirs:
        os.makedirs(d, exist_ok=True)


# ══════════════════════════════════════════════════════════
#  SECTION 3: DEBT WATERFALL CHART
#  Shows debt shrinking + equity growing year by year
#  This is the most intuitive LBO visual
# ══════════════════════════════════════════════════════════

def plot_debt_waterfall(debt_df: pd.DataFrame,
                        assumptions: DealAssumptions,
                        save_path: str = None):
    """
    Stacked bar chart showing capital structure evolution:
    - Debt (red) shrinks each year as FCF pays it down
    - Equity (green) grows as a share of total value
    - Entry EV line shows starting purchase price
    """
    years = list(debt_df.columns)

    opening_debt = debt_df.loc["Opening Debt (₹ Cr)"].tolist()
    fcf_used     = debt_df.loc["Debt Repayment (₹ Cr)"].tolist()
    closing_debt = debt_df.loc["Closing Debt (₹ Cr)"].tolist()

    # Implied equity value = Entry EV - Closing Debt each year
    # (simplified — in reality you'd mark-to-market EV each year)
    entry_ev     = assumptions.entry_ev
    equity_vals  = [round(entry_ev - d, 1) for d in closing_debt]

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    fig.patch.set_facecolor("#0f0f0f")

    for ax in axes:
        ax.set_facecolor("#1a1a1a")
        ax.tick_params(colors="white")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        for spine in ["bottom", "left"]:
            ax.spines[spine].set_color("#444")

    # ── Panel 1: Stacked bar — debt vs equity ─────────────
    ax1 = axes[0]
    x = np.arange(len(years))

    debt_bars   = ax1.bar(x, closing_debt, color="#e74c3c",
                          label="Remaining Debt", zorder=3,
                          edgecolor="#0f0f0f", linewidth=1.2)
    equity_bars = ax1.bar(x, equity_vals, bottom=closing_debt,
                          color="#2ecc71", label="Implied Equity",
                          zorder=3, edgecolor="#0f0f0f", linewidth=1.2)

    # Entry EV reference line
    ax1.axhline(y=entry_ev, color="#f39c12", linewidth=1.5,
                linestyle="--", label=f"Entry EV (₹{entry_ev:.0f}Cr)")

    # Annotate debt balance on each bar
    for bar, debt in zip(debt_bars, closing_debt):
        if debt > 20:
            ax1.text(bar.get_x() + bar.get_width() / 2,
                     debt / 2,
                     f"₹{debt:.0f}",
                     ha="center", va="center",
                     color="white", fontsize=9, fontweight="bold")

    ax1.set_title("Capital Structure Evolution\n(Debt Paydown Over Hold Period)",
                  color="white", fontsize=12, fontweight="bold", pad=10)
    ax1.set_ylabel("₹ Crore", color="#aaaaaa", fontsize=10)
    ax1.set_xticks(x)
    ax1.set_xticklabels([str(y) for y in years], color="white")
    ax1.legend(facecolor="#2a2a2a", edgecolor="#444",
               labelcolor="white", fontsize=9)
    ax1.grid(axis="y", color="#2a2a2a", zorder=0)

    # ── Panel 2: FCF bar chart — cash engine ──────────────
    ax2 = axes[1]
    fcf_vals = debt_df.loc["FCF Available (₹ Cr)"].tolist()
    leverage = debt_df.loc["Leverage Ratio (Debt/EBITDA)"].tolist()

    fcf_colors = ["#27ae60" if f > 0 else "#e74c3c" for f in fcf_vals]
    ax2.bar(x, fcf_vals, color=fcf_colors, zorder=3,
            edgecolor="#0f0f0f", linewidth=1.2, label="Annual FCF")

    # Annotate leverage ratio above each bar
    for i, (xi, lev) in enumerate(zip(x, leverage)):
        ax2.text(xi, fcf_vals[i] + 1.5,
                 f"{lev}x lev",
                 ha="center", va="bottom",
                 color="#aaaaaa", fontsize=8)

    ax2.set_title("Free Cash Flow Used for Debt Repayment\n(Leverage Ratio Declining)",
                  color="white", fontsize=12, fontweight="bold", pad=10)
    ax2.set_ylabel("₹ Crore (FCF)", color="#aaaaaa", fontsize=10)
    ax2.set_xticks(x)
    ax2.set_xticklabels([str(y) for y in years], color="white")
    ax2.grid(axis="y", color="#2a2a2a", zorder=0)

    fig.suptitle(f"LBO Debt Waterfall — {assumptions.company_name}",
                 color="white", fontsize=14, fontweight="bold", y=1.02)

    plt.tight_layout()

    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches="tight",
                    facecolor=fig.get_facecolor())
        print(f"  Chart saved -> {save_path}")

    plt.show()
    return fig


# ══════════════════════════════════════════════════════════
#  SECTION 4: EXCEL TEARSHEET EXPORT
#  Auto-generates a formatted multi-sheet Excel report
#  This is what you attach to your resume / send to recruiters
# ══════════════════════════════════════════════════════════

def export_excel_tearsheet(assumptions: DealAssumptions,
                            results: dict,
                            is_df: pd.DataFrame,
                            debt_df: pd.DataFrame,
                            scenario_table: pd.DataFrame,
                            heatmap1: pd.DataFrame,
                            heatmap2: pd.DataFrame,
                            save_path: str = "outputs/tearsheet/LBO_Tearsheet.xlsx"):
    """
    Export a 5-sheet Excel workbook:
    Sheet 1: Deal Summary
    Sheet 2: Income Statement
    Sheet 3: Debt Schedule
    Sheet 4: Scenario Comparison
    Sheet 5: Sensitivity Tables
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Color palette ──────────────────────────────────────
    C_DARK    = "0F0F0F"
    C_HEADER  = "1A1A2E"
    C_ACCENT  = "16213E"
    C_GREEN   = "2ECC71"
    C_RED     = "E74C3C"
    C_YELLOW  = "F1C40F"
    C_WHITE   = "FFFFFF"
    C_GREY    = "AAAAAA"
    C_BLUE    = "3498DB"

    thin = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_style(cell, bg=C_HEADER):
        cell.font      = Font(bold=True, color=C_WHITE, size=11)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    def data_style(cell, bold=False):
        cell.font      = Font(color=C_WHITE, bold=bold, size=10)
        cell.fill      = PatternFill("solid", fgColor="1E1E2E")
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border    = border

    def label_style(cell):
        cell.font      = Font(color=C_GREY, size=10)
        cell.fill      = PatternFill("solid", fgColor=C_ACCENT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = border

    # ══════════════════════════════════════════════════════
    # SHEET 1: Deal Summary
    # ══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Deal Summary"
    ws1.sheet_properties.tabColor = "2ECC71"
    ws1.sheet_view.showGridLines = False

    # Title block
    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value     = f"LBO INVESTMENT TEARSHEET — {assumptions.company_name.upper()}"
    title_cell.font      = Font(bold=True, color=C_WHITE, size=14)
    title_cell.fill      = PatternFill("solid", fgColor="16213E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Summary data
    summary_data = [
        ("DEAL STRUCTURE", ""),
        ("Entry EV/EBITDA Multiple", f"{assumptions.entry_ev_multiple}x"),
        ("Entry Enterprise Value",   f"₹{assumptions.entry_ev:,.0f} Cr"),
        ("Total Debt Raised",        f"₹{assumptions.total_debt:,.0f} Cr ({assumptions.debt_pct*100:.0f}%)"),
        ("Equity Invested",          f"₹{assumptions.equity_invested:,.0f} Cr ({(1-assumptions.debt_pct)*100:.0f}%)"),
        ("Interest Rate",            f"{assumptions.interest_rate*100:.1f}%"),
        ("Holding Period",           f"{assumptions.holding_period} Years"),
        ("", ""),
        ("EXIT & RETURNS", ""),
        ("Exit EV/EBITDA Multiple",  f"{assumptions.exit_ev_multiple}x"),
        ("Exit Enterprise Value",    f"₹{results['exit_ev']:,.0f} Cr"),
        ("Remaining Debt at Exit",   f"₹{results['exit_debt']:,.0f} Cr"),
        ("Exit Equity Value",        f"₹{results['exit_equity']:,.0f} Cr"),
        ("IRR",                      f"{results['irr']*100:.1f}%"),
        ("Money-on-Money Multiple",  f"{results['mom']:.2f}x"),
        ("", ""),
        ("OPERATING ASSUMPTIONS", ""),
        ("Base Year Revenue",        f"₹{assumptions.revenue_base:,.0f} Cr"),
        ("Revenue CAGR",             f"{assumptions.revenue_cagr*100:.0f}%"),
        ("EBITDA Margin",            f"{assumptions.ebitda_margin*100:.0f}%"),
        ("Tax Rate",                 f"{assumptions.tax_rate*100:.0f}%"),
        ("Capex % Revenue",          f"{assumptions.capex_pct_revenue*100:.0f}%"),
    ]

    for i, (label, value) in enumerate(summary_data, start=3):
        r = i
        if label in ("DEAL STRUCTURE", "EXIT & RETURNS", "OPERATING ASSUMPTIONS"):
            ws1.merge_cells(f"A{r}:F{r}")
            c = ws1[f"A{r}"]
            c.value     = label
            c.font      = Font(bold=True, color=C_WHITE, size=11)
            c.fill      = PatternFill("solid", fgColor="16213E")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws1.row_dimensions[r].height = 20
        elif label == "":
            ws1.row_dimensions[r].height = 8
        else:
            lc = ws1[f"A{r}"]
            vc = ws1[f"B{r}"]
            lc.value = label
            vc.value = value
            label_style(lc)
            data_style(vc, bold=(label in ("IRR", "Money-on-Money Multiple")))

            # Color IRR cell based on performance
            if label == "IRR":
                irr_val = results["irr"] * 100
                color = C_GREEN if irr_val >= 20 else (C_YELLOW if irr_val >= 15 else C_RED)
                vc.font = Font(bold=True, color=color, size=12)

            ws1.row_dimensions[r].height = 18

    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22

    # Return attribution block
    row_start = len(summary_data) + 5
    ws1.merge_cells(f"A{row_start}:F{row_start}")
    c = ws1[f"A{row_start}"]
    c.value     = "RETURN ATTRIBUTION"
    c.font      = Font(bold=True, color=C_WHITE, size=11)
    c.fill      = PatternFill("solid", fgColor="16213E")
    c.alignment = Alignment(horizontal="left")

    for j, (k, v) in enumerate(results["attribution"].items(), start=row_start+1):
        lc = ws1[f"A{j}"]
        vc = ws1[f"B{j}"]
        lc.value = k
        vc.value = f"₹{v:,.1f} Cr"
        label_style(lc)
        data_style(vc)
        ws1.row_dimensions[j].height = 18

    # ══════════════════════════════════════════════════════
    # SHEET 2: Income Statement
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Income Statement")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "3498DB"

    ws2.merge_cells("A1:G1")
    t = ws2["A1"]
    t.value     = "PROJECTED INCOME STATEMENT & FREE CASH FLOW"
    t.font      = Font(bold=True, color=C_WHITE, size=13)
    t.fill      = PatternFill("solid", fgColor="16213E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # Headers
    years = list(is_df.columns)
    ws2["A2"].value = "Line Item"
    header_style(ws2["A2"])
    for col_i, yr in enumerate(years, start=2):
        c = ws2.cell(row=2, column=col_i, value=str(yr))
        header_style(c)
    ws2.row_dimensions[2].height = 20

    # FCF divider rows
    fcf_rows = {"Free Cash Flow (₹ Cr)", "(-) Capex (₹ Cr)",
                "(-) ΔWorking Capital (₹ Cr)", "D&A (₹ Cr)"}

    for row_i, (idx, row_data) in enumerate(is_df.iterrows(), start=3):
        lc = ws2.cell(row=row_i, column=1, value=idx)
        label_style(lc)

        is_fcf_section = idx in fcf_rows

        for col_i, val in enumerate(row_data, start=2):
            c = ws2.cell(row=row_i, column=col_i, value=val)
            data_style(c, bold=(idx in ("EBITDA (₹ Cr)", "Free Cash Flow (₹ Cr)")))

            if idx == "Free Cash Flow (₹ Cr)":
                color = C_GREEN if val >= 0 else C_RED
                c.font = Font(bold=True, color=color, size=10)

        ws2.row_dimensions[row_i].height = 18

    ws2.column_dimensions["A"].width = 30
    for col_i in range(2, len(years) + 2):
        ws2.column_dimensions[get_column_letter(col_i)].width = 14

    # ══════════════════════════════════════════════════════
    # SHEET 3: Debt Schedule
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Debt Schedule")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "E74C3C"

    ws3.merge_cells("A1:G1")
    t = ws3["A1"]
    t.value     = "DEBT REPAYMENT SCHEDULE"
    t.font      = Font(bold=True, color=C_WHITE, size=13)
    t.fill      = PatternFill("solid", fgColor="16213E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    ws3["A2"].value = "Line Item"
    header_style(ws3["A2"])
    for col_i, yr in enumerate(list(debt_df.columns), start=2):
        c = ws3.cell(row=2, column=col_i, value=str(yr))
        header_style(c)

    for row_i, (idx, row_data) in enumerate(debt_df.iterrows(), start=3):
        lc = ws3.cell(row=row_i, column=1, value=idx)
        label_style(lc)
        for col_i, val in enumerate(row_data, start=2):
            c = ws3.cell(row=row_i, column=col_i, value=val)
            data_style(c, bold=("Closing Debt" in idx or "FCF" in idx))
            if "Closing Debt" in idx:
                c.font = Font(bold=True, color=C_GREEN, size=10)
        ws3.row_dimensions[row_i].height = 18

    ws3.column_dimensions["A"].width = 30
    for col_i in range(2, len(list(debt_df.columns)) + 2):
        ws3.column_dimensions[get_column_letter(col_i)].width = 16

    # ══════════════════════════════════════════════════════
    # SHEET 4: Scenario Analysis
    # ══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Scenario Analysis")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = "F1C40F"

    ws4.merge_cells("A1:K1")
    t = ws4["A1"]
    t.value     = "SCENARIO ANALYSIS — BULL / BASE / BEAR / DISTRESSED"
    t.font      = Font(bold=True, color=C_WHITE, size=13)
    t.fill      = PatternFill("solid", fgColor="16213E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    # Header row
    cols = ["Scenario"] + list(scenario_table.columns)
    for col_i, col_name in enumerate(cols, start=1):
        c = ws4.cell(row=2, column=col_i, value=col_name)
        header_style(c)

    # Data rows with scenario-based coloring
    scenario_colors = {
        "Bull": "1A5C38", "Base": "16213E",
        "Bear": "5C3A1A", "Distressed": "5C1A1A"
    }

    for row_i, (scenario_name, row_data) in enumerate(scenario_table.iterrows(), start=3):
        bg = scenario_colors.get(scenario_name.split()[-1], "1E1E2E")
        c = ws4.cell(row=row_i, column=1, value=scenario_name)
        c.font  = Font(bold=True, color=C_WHITE, size=10)
        c.fill  = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left")
        c.border = border

        for col_i, val in enumerate(row_data, start=2):
            cell = ws4.cell(row=row_i, column=col_i, value=val)
            cell.font      = Font(color=C_WHITE, size=10,
                                   bold=(col_i >= len(cols) - 1))
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border
        ws4.row_dimensions[row_i].height = 20

    for col_i in range(1, len(cols) + 1):
        ws4.column_dimensions[get_column_letter(col_i)].width = 18

    # ══════════════════════════════════════════════════════
    # SHEET 5: Sensitivity Tables
    # ══════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Sensitivity")
    ws5.sheet_view.showGridLines = False
    ws5.sheet_properties.tabColor = "9B59B6"

    def write_heatmap_to_sheet(ws, df, start_row, title):
        """Write a sensitivity table with color-coded IRR cells"""
        # Title
        end_col = get_column_letter(len(df.columns) + 1)
        ws.merge_cells(f"A{start_row}:{end_col}{start_row}")
        t = ws[f"A{start_row}"]
        t.value     = title
        t.font      = Font(bold=True, color=C_WHITE, size=12)
        t.fill      = PatternFill("solid", fgColor="16213E")
        t.alignment = Alignment(horizontal="center")
        ws.row_dimensions[start_row].height = 22

        # Axis label
        ws.cell(row=start_row+1, column=1, value=df.index.name).font = \
            Font(italic=True, color=C_GREY, size=9)

        # Column headers
        ws.cell(row=start_row+1, column=1, value=f"↓ {df.index.name}  |  {df.columns.name} →")
        for col_i, col_name in enumerate(df.columns, start=2):
            c = ws.cell(row=start_row+1, column=col_i, value=col_name)
            header_style(c)

        # Data with IRR-based coloring
        for row_i, (idx, row_data) in enumerate(df.iterrows(), start=start_row+2):
            lc = ws.cell(row=row_i, column=1, value=idx)
            label_style(lc)

            for col_i, val in enumerate(row_data, start=2):
                c = ws.cell(row=row_i, column=col_i, value=val)

                if pd.isna(val):
                    c.value = "N/A"
                    bg_color = "333333"
                elif val >= 25:
                    bg_color = "1A5C38"   # dark green — exceptional
                elif val >= 20:
                    bg_color = "27AE60"   # green — good
                elif val >= 15:
                    bg_color = "D4AC0D"   # yellow — marginal
                elif val >= 10:
                    bg_color = "CA6F1E"   # orange — weak
                else:
                    bg_color = "922B21"   # red — avoid

                c.font      = Font(bold=True, color=C_WHITE, size=10)
                c.fill      = PatternFill("solid", fgColor=bg_color)
                c.alignment = Alignment(horizontal="center")
                c.border    = border

            ws.row_dimensions[row_i].height = 18

        for col_i in range(1, len(df.columns) + 2):
            ws.column_dimensions[get_column_letter(col_i)].width = 13

        return start_row + len(df) + 4

    next_row = write_heatmap_to_sheet(
        ws5, heatmap1, 1,
        "SENSITIVITY: Entry Multiple × Exit Multiple → IRR (%)"
    )
    write_heatmap_to_sheet(
        ws5, heatmap2, next_row,
        "SENSITIVITY: Revenue CAGR × EBITDA Margin → IRR (%)"
    )

    # ── Save workbook ──────────────────────────────────────
    wb.save(save_path)
    print(f"\n  Excel tearsheet saved -> {save_path}")
    return save_path


# ══════════════════════════════════════════════════════════
#  SECTION 5: MAIN RUNNER
# ══════════════════════════════════════════════════════════

def main():
    print("\n" + "█" * 60)
    print("  LBO FINANCIAL MODEL — PYTHON")
    print("  Built for PE/VC Interview Preparation")
    print("█" * 60)

    # ── Parse CLI args ─────────────────────────────────────
    args = parse_args()

    # ── Build assumptions ──────────────────────────────────
    deal = DealAssumptions()
    if args.company : deal.company_name     = args.company
    if args.entry   : deal.entry_ev_multiple = args.entry
    if args.exit    : deal.exit_ev_multiple  = args.exit
    if args.cagr    : deal.revenue_cagr      = args.cagr
    if args.debt    : deal.debt_pct          = args.debt
    if args.hold    : deal.holding_period    = args.hold

    deal.summary()
    setup_directories()

    # ── Step 1: Returns (runs IS + debt schedule internally) ──
    print("\n  [1/5] Calculating base case returns...")
    results = calculate_returns(deal)
    print_returns(results, deal)

    debt_df = results["debt_df"]
    is_df   = results["is_df"]

    print("\n  [2/5] Income Statement:")
    print_income_statement(is_df)

    print("\n  [3/5] Debt Schedule:")
    print_debt_schedule(debt_df)

    # ── Step 2: Scenarios ──────────────────────────────────
    print("\n  [4/5] Scenario Analysis...")
    all_scenarios  = run_all_scenarios(deal)
    scenario_table = build_scenario_table(all_scenarios)

    print("\n  SCENARIO TABLE:")
    print("  " + "=" * 80)
    print(scenario_table.to_string())

    # ── Step 3: Charts ────────────────────────────────────
    if not args.no_plots:
        print("\n  [5/5] Generating charts...")

        plot_debt_waterfall(
            debt_df, deal,
            save_path="outputs/charts/debt_waterfall.png"
        )
        plot_scenario_comparison(
            all_scenarios,
            save_path="outputs/charts/scenarios.png"
        )
        h1, h2 = run_all_sensitivities(
            deal,
            save_dir="outputs/charts/"
        )
    else:
        print("\n  [5/5] Skipping charts (--no-plots flag set)")
        from sensitivity import run_sensitivity
        h1 = run_sensitivity(deal, "entry_ev_multiple",
                             [6,7,8,9,10,11],
                             "exit_ev_multiple", [7,8,9,10,11])
        h2 = run_sensitivity(deal, "revenue_cagr",
                             [0.04,0.06,0.08,0.10,0.12,0.15,0.18],
                             "ebitda_margin",
                             [0.12,0.14,0.16,0.18,0.20,0.22])

    # ── Step 4: Export Excel tearsheet ────────────────────
    print("\n  Exporting Excel tearsheet...")
    export_excel_tearsheet(
        assumptions     = deal,
        results         = results,
        is_df           = is_df,
        debt_df         = debt_df,
        scenario_table  = scenario_table,
        heatmap1        = h1,
        heatmap2        = h2,
        save_path       = "outputs/tearsheet/LBO_Tearsheet.xlsx"
    )

    # ── Final summary ──────────────────────────────────────
    print("\n" + "█" * 60)
    print(f"  COMPLETE. Results for {deal.company_name}")
    print(f"  Base IRR   : {results['irr']*100:.1f}%")
    print(f"  Base MoM   : {results['mom']:.2f}x")
    print(f"  Charts     : outputs/charts/")
    print(f"  Tearsheet  : outputs/tearsheet/LBO_Tearsheet.xlsx")
    print("█" * 60 + "\n")


if __name__ == "__main__":
    main()
